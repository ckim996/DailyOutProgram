from datetime import date, datetime
import pandas as pd
from src.shipstation.client import get_shipments
import os
from src.lookup.sku_lookup import lookup_sku
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import Counter, defaultdict
from generate_test_data import generate_test_orders
from src.shipping.engine import get_carrier_service
from src.shipping.engine import get_sku_info_from_dailyouttools, get_weight_from_pkg_string
from src.shipping.optimizer import shop_and_optimize
from src.shipstation.rates import get_live_rates
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
import config
import re

base_font = Font(size=9)
store_font = Font(size=12,bold=True)

WRAP_COLUMNS = {"Part#", "Interchange #", "Attention"}

UPS_STATES = {
    # Abbreviations
    "AZ", "CA", "CO", "IA", "ID", "IL", "KS", "LA", "MN", "MO", "MS",
    "MT", "NE", "NM", "NV", "OK", "OR", "SD", "TX", "UT", "WA", "WI", "WY",
    # Full names
    "Arizona", "California", "Colorado", "Iowa", "Idaho", "Illinois", "Kansas",
    "Louisiana", "Minnesota", "Missouri", "Mississipi", "Montana", "Nebraska",
    "New Mexico", "Nevada", "Oklahoma", "Oregon", "South Dakota", "Texas",
    "Utah", "Washington", "Wisconsin", "Wyoming"
}

UPS_STATES_NORMALIZED = {s.strip().lower() for s in UPS_STATES}

HEADERS = [
    "Sequence", "Order #", "SKU", "Part#", "Interchange #",
    "Qty", "Carrier", "Service", "Box", "Shipping Price", "Attention"
]

thin = Side(style="thin")
all_border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin
)

def get_store_name(store_id):
    """
        Converts the store_id into Actual Platform Store Name
        Args: store_id: retrieved from Shipstation API
        Returns: The name of the store
        Used: main.py
    """
    if store_id is None:
        return "Unknown Store"
    return config.STORE_MAP.get(store_id, f"Store {store_id}")

def count_awaiting_shipments():
    """
        Displays the number of orders in awaiting shipment from shipstation
        Returns: # of orders
        Used: app.py
    """
    shipments = get_shipments() 
    awaiting = [s for s in shipments if s['orderStatus'] == 'awaiting_shipment'] 
    return len(awaiting)

# Method to help determine if an order has ebay Purchase GP#
def load_lp_data():
    """
        Reads DailyOutTools.xlsx (LP sheet) and find GP#s out of stock to flag them for ebay purchase
        Returns: lp_map dictionary GP#s and its LP Price
        Used: main.py
    """
    lp_path = "data/DailyOuttools.xlsx"
    try:
        df_lp = pd.read_excel(lp_path, sheet_name="LP", usecols="B:D")
        lp_map = {}
        for _, row in df_lp.iterrows():
            gp = str(row['Omni5']).strip()
            qoh = row.get('QOH', 0)
            price = row.get('L/P', 0)
            if qoh == 0 and price > 0:
                lp_map[gp] = price
        return lp_map
    except Exception as e:
        print(f"LP lookup error: {e}")
        return {}
    
def create_list_algorithm(wb, parts_list):

    """
        Generates "List Algorithm" sheet and organizes GP#s into appropriate printable format.
        Appropriate format: follows the 40 -> 41 -> 40 -> 41 ... depending on the number of rows
        :param parts_list: The raw list of every gp# before being comma-delimited
        Used: main.py -> app.py
    """

    if "List Algorithm" in wb.sheetnames:
        del wb["List Algorithm"]

    ws = wb.create_sheet("List Algorithm")
    ws.views.sheetView[0].view = 'pageLayout'

    # This makes Excel show the full page width in Page Layout view
    widths = {'A': 18, 'B': 5, 'C': 2, 'D': 18, 'E': 5, 'F': 2, 'G': 18, 'H': 5, 'I': 2}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Determine total pages needed based on 40/41 pattern
    total_part_rows = len(parts_list)
    temp_count = total_part_rows
    pages_to_run = 0
    while temp_count > 0:
        pages_to_run += 1
        limit = 40 if pages_to_run % 2 != 0 else 41
        temp_count -= limit

    current_part_idx = 0
    page_offset = 0
    MAX_V_ROWS = 44
    COL_GROUPS = [(1,2),(4,5),(7,8)]

    for p_num in range(1, pages_to_run + 1):
        page_limit = 40 if p_num % 2 != 0 else 41
        raw_slice = parts_list[current_part_idx : current_part_idx + page_limit]
        current_part_idx += page_limit

        # process data for current page
        page_gps= []
        for raw_str in raw_slice:
            split_parts = [p for p in re.split(r"[,\s]+", str(raw_str)) if p]
            page_gps.extend(split_parts)

        counts = Counter(page_gps)
        sorted_gps = sorted(counts.items())

        anchor_row = page_offset + MAX_V_ROWS
        ws.cell(row=anchor_row, column=9, value="")

        # Write to the grid: Fill row 1-44 for col A-B, D-E, G-H
        sorted_idx = 0
        for col_gp, col_qty in COL_GROUPS:
            for r in range(1, MAX_V_ROWS + 1):
                if sorted_idx < len(sorted_gps):
                    gp_name, qty = sorted_gps[sorted_idx]
                    target_row = page_offset + r

                    ws.cell(row=target_row, column=col_gp, value=gp_name).font = Font(bold=True)
                    ws.cell(row=target_row, column=col_qty, value=qty).alignment = Alignment(horizontal="center")

                    sorted_idx += 1

            # If we run out of sorted GPs for this page, stop filling column groups
            if sorted_idx >= len(sorted_gps):
                break

        # Move page_offset for the next physical page
        page_offset += MAX_V_ROWS

    ws.print_area = f'A1:I{page_offset}'

def fetch_order_data(row, order_total_qty, sku_info, lp_lookup):
    """
        extract_todays_shipments -> write_grouped_excel -> fetch_order_data
        Method runs in parallel to calculate shipping costs, determine the best carrier,
        check if item is ebay purchase, and calculate potential savings
        
        :param row: row is the store_row extracted from extract_todays_shipments [values are in extract_todays_shipments]
        :param sku_info: SKU row coming from DailyOutTools; [values are the headers from the sheet 'DB' or 'Nonmounts' from DailyOutTools]
        :param lp_lookup: GP#s dictionary that has GP# and its LP Price coming from 'LP' Sheet inside DailyOutTools
    """
    order_no = row.get("Order #")
    shippingDB_cost = float(sku_info.get("Shipping DB", 0) or 0) if sku_info else 0.0
    total_qty = order_total_qty.get(order_no, 0)
    
    # LP Logic
    lp_found_info = []
    is_ebay_purchase = False
    if sku_info:
        raw_parts = str(sku_info.get("Part #", "") or "")
        parts_list = [p.strip() for p in raw_parts.split(",") if p.strip()]
        for p in parts_list:
            if p in lp_lookup:
                lp_found_info.append(f"{p} (${lp_lookup[p]})")
                is_ebay_purchase = True

    # Decision Logic
    try:
        decision = get_carrier_service(row)
    except Exception as e:
        decision = None
    best_rate = None
    best_rate_cost = 0.0
    full_service_display = "N/A"
    decision_msg = ""
    dims = None
    savings = 0.0

    print(f"Decision({order_no}): {decision}")

    if total_qty > 1:
        decision_msg = "NEED WAREHOUSE ASSISTANCE"
        c, s, p, w, dims = (None, None, None, None, None)
    elif decision is None:
        decision_msg = "ERROR: No Decision Generated"
        c, s, p, w, dims = (None, None, None, None, None)
    elif decision[0] == "ERROR":
        decision_msg = decision[1]
        c, s, p, w, dims = (None, None, None, None, None)
    else:
        c, s, p, w, dims = decision
        if c != "SHOP_RATES":
            rate_results, _ = get_live_rates(order_no, c, s, p, w, dims, row.get("State"), row.get("Zip"),is_residential=False)
            if rate_results:
                best_rate = rate_results[0]
                raw_pkg = best_rate.get("packageType")
                correct_pkg = next((k for k,v in config.pkg_map.items() if v == raw_pkg), raw_pkg)
                best_rate["winning_pkg_str"] = correct_pkg
                s_name = best_rate.get("serviceName") or s
                best_rate.update({
                    "serviceName": s_name,
                    "comparison_log": f"{s_name} only"
                })
        else:
            store_id = row.get("Store")
            best_rate = shop_and_optimize(order_no, w, dims, row.get("State"), row.get("Zip"), sku_info,store_id=store_id, is_residential=False)
        
        best_rate_cost = best_rate.get("shipmentCost", 0.0) if best_rate else 0.0
        
        if best_rate:
            
            carrier_val, service_val, box_val = "N/A", "N/A", "N/A"
            
            raw_carrier = str(best_rate.get("carrierCode")).lower()
            raw_service = str(best_rate.get("serviceCode")).lower()
            raw_pkg = str(best_rate.get("packageType", "")).lower()

            if best_rate.get("winning_pkg_str"):
                p = best_rate.get("winning_pkg_str")
            
            # Map Carriers for Excel
            if "ups" in raw_carrier:
                carrier_val, service_val, box_val = "U", "", p
            elif "usps" in raw_carrier or "stamps" in raw_carrier:
                carrier_val = "P"
                if "usps_ground_advantage" in raw_service: service_val, box_val = "G", p
                elif "usps_first_class_mail" in raw_service: service_val, box_val = "F", "B"
                else:
                    service_val = "P"
                    if "flat_rate_padded_envelope" in raw_pkg: box_val = "P"
                    elif "flat_rate_envelope" in raw_pkg: box_val = "F"
                    elif "medium_flat_rate_box" in raw_pkg: box_val = "M"
                    elif "large_flat_rate_box" in raw_pkg: box_val = "L"
                    else: box_val = p

            current_weight = sku_info.get("Weight") if sku_info else None
            if box_val == "B" and sku_info:
                pkg_string = sku_info.get("Package","")
                rule_3_weight = get_weight_from_pkg_string(pkg_string)

                if rule_3_weight is not None:
                    current_weight = rule_3_weight
            
            full_service_display = f"{best_rate.get('serviceName')} (${best_rate_cost})"
            savings = round(shippingDB_cost - best_rate_cost, 2) if best_rate_cost > 0 else 0.0
            decision_msg = savings if best_rate_cost > 0 else "Rate Fetch Error"
            store_id = row.get("Store")

            if best_rate.get("winning_pkg_str"):
                print(f"{order_no} | winning_pkg_str: {best_rate.get("winning_pkg_str")}")

            log_entry = {
                "Order #": order_no,
                "SKU": row.get("SKU"),
                "DB Cost": shippingDB_cost,
                "Winner": full_service_display,
                "Savings": savings,
                "Decision Type": best_rate.get("serviceCode", "N/A"),
                "Pkg": best_rate.get("winning_pkg_str",""),
                "Comparison": best_rate.get("comparison_log", "N/A"),
                "Delivery Time": f"{(date.fromisoformat(best_rate.get('estimated_delivery_date')[:10]) - date.today()).days} Days" if best_rate.get('estimated_delivery_date') else "N/A",
                "Arrival": best_rate.get("estimated_delivery_date", "N/A")[:10],
                "Fallback": "FALLBACK" if best_rate.get("is_priority_fallback") else "",
                "LP": ", ".join(lp_found_info),
                "Weight": current_weight,
                "Dims": f"{int(dims[0])}x{int(dims[1])}x{int(dims[2])}" if dims else "",
                "Shipping Cost":best_rate_cost,
                "GP": sku_info.get("Part #","") if sku_info else "",
                "Interchange": sku_info.get("Interchange (not in order)","") if sku_info else "",
                "Store Name":get_store_name(row.get("Store")),
                "Shipping Status":""
            }

            return {
                "order_no": order_no,
                "best_rate_cost": best_rate_cost,
                "decision_msg": decision_msg,
                "savings": savings,
                "is_ebay": is_ebay_purchase,
                "lp_info": ", ".join(lp_found_info),
                "excel_mapping": {"Carrier": carrier_val, "Service": service_val, "Box": box_val},
                "log_entry": log_entry,
                "decision": decision
            }
    
    return {
        "order_no": order_no, 
        "decision_msg": decision_msg, 
        "is_ebay": is_ebay_purchase, 
        "best_rate": None,
        "excel_mapping": {}, # Keeps the main loop from crashing
        "savings": 0.0
    }

progress_status = {"percent": 0}
def write_grouped_excel(store_rows, output_file):

    """
        extract_todays_shipments -> write_grouped_excel -> fetch_order_data
        gets store_rows from extract_todays_shipments

        Using the processed data from store_rows, this method builds the final Excel Workbook.
        Manages complex formatting, merge cells for multi-item orders, creates the Decision Log Sheet as well.
        Used: main.py
    """

    global progress_status
    progress_status = {"percent": 0}
    today_day = str(datetime.today().day)
    
    try:
        wb = load_workbook(output_file)
    except FileNotFoundError:
        print(f"Error: {output_file} not found.")
        return

    ### CHANGE: Verify template exists before proceeding
    if 'Copy' not in wb.sheetnames:
        print("Error: Template sheet 'Copy' not found.")
        return

    ### CHANGE: Scrub the 'Copy' template of any old data/merges below headers
    template = wb['Copy']
    template.delete_rows(3, template.max_row)

    ### CHANGE: Remove existing daily sheet to avoid naming conflicts (e.g., '19 1')
    if today_day in wb.sheetnames:
        del wb[today_day]

    ### CHANGE: Create daily sheet as a direct clone of the clean template
    ws = wb.copy_worksheet(template)
    ws.title = today_day

    # Header Formatting for Daily Sheet
    header_font = Font(size=8, color="FFFFFF")
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    for col_idx, col_name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = black_fill
        cell.alignment = Alignment(horizontal="center")
        ws.cell(row=2, column=col_idx).fill = black_fill

    # Styles (Defining these once for performance)
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    # Pre-compute totals for highlighting
    customer_counter = Counter()
    order_total_qty = defaultdict(int)
    unique_rows_to_fetch = []
    seen_orders = set()

    # Check duplicate customers or orders with qty > 1
    for store_rows_list in store_rows.values():
        for row in store_rows_list:
            first = str(row.get("First Name", "")).strip().lower()
            last = str(row.get("Last Name", "")).strip().lower()
            if first and last:
                customer_counter[(first, last)] += 1
            
            order_no = row.get("Order #")
            try:
                qty = int(row.get("Qty", 0))
                order_total_qty[order_no] += qty
            except (TypeError, ValueError):
                pass
            if order_no not in seen_orders:
                unique_rows_to_fetch.append(row)
                seen_orders.add(order_no)

    # reads the DailyOutTools to find possible ebay purchase GP#s
    lp_lookup = load_lp_data()

    # Parallel Fetching Orders
    total_unique = len(unique_rows_to_fetch)
    rate_results_map = {}
    with ThreadPoolExecutor(max_workers=5) as executor:
        # Use submit instead of map to track individual completions
        future_to_order = {
            executor.submit(fetch_order_data, r, order_total_qty, get_sku_info_from_dailyouttools(r.get("SKU")), lp_lookup): r.get("Order #") 
            for r in unique_rows_to_fetch
        }
        
        completed_fetch = 0
        for future in as_completed(future_to_order):
            res = future.result()
            rate_results_map[res["order_no"]] = res
            
            completed_fetch += 1
            progress_status["percent"] = int((completed_fetch / total_unique) * 100)

    current_row = 3
    grand_total_savings = 0.0
    decision_logs = []
    all_parts_for_list = []

    for store_id in config.STORE_MAP:
        if store_id not in store_rows:
            continue

        # Sorting Order # per Store
        def order_sort_key(row):
            order_val = str(row.get("Order #", "0"))
            numeric_only = order_val.replace("-","")
            try:
                return int(numeric_only)
            except ValueError:
                return 0
        store_rows[store_id].sort(key=order_sort_key)

        rows = store_rows[store_id]
        store_name = get_store_name(store_id)
        
        ### CHANGE: New trackers to handle Sequence-by-Order and Store-specific merging
        store_order_tracker = {}
        order_sequence_map = {}  
        next_seq_num = 1         
        
        ws.cell(row=current_row, column=1, value=store_name).font = store_font
        current_row += 1

        for row in rows:
            order_no = row.get("Order #")
            sku_info = get_sku_info_from_dailyouttools(row.get("SKU"))

            # Get pre-fetched data from our parallel map
            res = rate_results_map.get(order_no,{})

            # Track start/end rows for merging this specific order
            if order_no not in store_order_tracker:
                store_order_tracker[order_no] = {"start": current_row, "end": current_row}
                is_first_row = True
            else:
                store_order_tracker[order_no]["end"] = current_row
                is_first_row = False

            # Get Shipping DB from DailyOutTools
            shippingDB_cost = 0.0
            if sku_info:
                shippingDB_cost = float(sku_info.get("Shipping DB", 0) or 0)
                print("----------------------------------------------------------------------------------------------")
                print(f"Order#:{order_no} | SKU Info: {sku_info} : shipping DB: {shippingDB_cost}")

            # --- GET PRE-COMPUTED LP DATA ---
            is_ebay_purchase = res.get("is_ebay", False)
            # IMPORTANT: We still need to collect parts for the final list algorithm
            if sku_info:
                raw_parts = str(sku_info.get("Part #", "") or "")
                if raw_parts and raw_parts.lower() != "nan":
                    all_parts_for_list.append(raw_parts)

            if res.get("excel_mapping"):
                row.update(res["excel_mapping"])

            best_rate_cost = res.get("best_rate_cost", 0)
            row["Shipping Price"] = best_rate_cost
            
            # Only the first row of an order gets the savings message and log entry
            if is_first_row:
                decision_msg = res.get("decision_msg", "")
                grand_total_savings += res.get("savings", 0.0)
                
                if "log_entry" in res:
                    raw_pkg = res["log_entry"].get("Pkg","")
                    # res["log_entry"]["Pkg"] = BOX_MAP.get(raw_pkg, raw_pkg)
                    decision_logs.append(res["log_entry"])
            else:
                decision_msg = ""

            # Write "possibly ebay purchase" to the daily sheet column 14
            if is_ebay_purchase:
                ebay_cell = ws.cell(row=current_row, column=14, value="possibly ebay purchase")
                ebay_cell.fill = yellow_fill
                ebay_cell.font = Font(size=9, bold=True)
                ebay_cell.alignment = Alignment(horizontal="left")

            # SKU column 15 (weight) and 16 (service)
            if sku_info:
                sku_weight = sku_info.get("Weight", 0)
                weight_cell = ws.cell(row=current_row, column=15, value=sku_weight)
                weight_cell.font = base_font
                weight_cell.alignment = Alignment(horizontal="center")

                sku_service = res.get("log_entry", {}).get("Decision Type", "N/A")
                service_cell = ws.cell(row=current_row, column=16, value=sku_service)
                service_cell.font = base_font
                service_cell.alignment = Alignment(horizontal="left")
            
            ### CHANGE: Logic to ensure Sequence matches the Order #, not the Row count
            if order_no not in order_sequence_map:
                order_sequence_map[order_no] = next_seq_num
                next_seq_num += 1
            
            row["Sequence"] = order_sequence_map[order_no]

            decision = res.get("decision")  # This is the (c, s, p, w, dims) tuple
            print(f"Decision for order({order_no}: {decision})")
            if decision and isinstance(decision, (list, tuple)) and len(decision) >= 5:
                c, s, p, w, dims = decision
            else:
                c, s, p, w, dims = (None, None, None, None, None)

            if order_no:
                print(f"\n[DEBUG] Order: {order_no}")
                print(f"  Raw Decision C: '{c}' | S: '{s}'")
                print(f"  Normalized Carrier: '{str(c or '').strip().lower()}'")
                print(f"  Normalized Service: '{str(s or '').strip().lower()}'")
                print(f"  State: '{row.get('State')}' (In List: {str(row.get('State')).strip().lower() in UPS_STATES_NORMALIZED})")
            # --------------------
            
            ex_map = res.get("excel_mapping", {})
            f_carrier = str(ex_map.get("Carrier", "")).strip().upper()
            f_service = str(ex_map.get("Service", "")).strip().upper()
            f_box = str(ex_map.get("Box", "")).strip()

            row.update({"Carrier": f_carrier, "Service": f_service, "Box": f_box})

            # Write data to cells
            for col_idx, col_name in enumerate(HEADERS, start=1):
                
                value = row.get(col_name)

                is_usps_priority = (f_carrier == "P" and f_service == "P")
                is_usps_first_class = (f_box == "B")

                if col_name == "Box":

                    if is_usps_priority and f_box in ["L", "M", "F", "P"]:
                        value = f_box
                        print(f"  [DEBUG] Preserving Flat Rate: {f_box}")
                    else:
                        mapped_box = config.BOX_MAP.get(f_box, f_box)
                        value = mapped_box
                        if f_box != mapped_box:
                            print(f"  [DEBUG] Ground/UPS detected. Mapping '{f_box}' -> '{mapped_box}'")
                        else:
                            print(f"  [DEBUG] No mapping found for '{f_box}', using literal.")
                            
                elif col_name == "Attention":
                    state = str(row.get("State", "")).strip().lower()
                    current_attn = str(row.get("Attention","") or "").strip()

                    if state in UPS_STATES_NORMALIZED and not (f_carrier == "U" or f_service == "G"):
                        value = f"ups | {current_attn}".strip(" | ")
                    else:
                        value = current_attn

                    if not(is_usps_priority or is_usps_first_class):
                        if dims and all(v > 1 for v in dims):
                            dim_str = f"{int(dims[0])}x{int(dims[1])}x{int(dims[2])}"
                            value = f"{value} | {dim_str}".strip(" | ")
                    
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.font = base_font
                cell.border = all_border

                if col_name in ["Sequence", "Carrier", "Service", "Box", "Shipping Price", "Qty"]:
                    cell.alignment = Alignment(horizontal="center")

                if col_name in WRAP_COLUMNS:
                    cell.alignment = Alignment(
                        wrap_text=True,
                        horizontal="left",
                        vertical="center"
                        )
                    ws. row_dimensions[current_row].height = None

                if col_name == "Part#" and value in (None, "", "None"):
                    cell.fill = red_fill

                customer_key = (
                    str(row.get("First Name", "")).strip().lower(),
                    str(row.get("Last Name", "")).strip().lower()
                )
                if col_name == "Order #" and customer_counter.get(customer_key, 0) > 1:
                    cell.fill = yellow_fill

                if col_name == "Qty" and order_total_qty.get(order_no, 0) > 1:
                    cell.fill = blue_fill

                if col_name == "Order #":
                    cell.alignment = Alignment(horizontal="right")

            res_cell = ws.cell(row=current_row, column=13, value=decision_msg)
            res_cell.font = base_font
            res_cell.border = all_border

            if isinstance(decision_msg, (int,float)):
                res_cell.alignment = Alignment(horizontal="center")
                res_cell.number_format = '"$"#,##0.00'
                if decision_msg > 0:
                    res_cell.fill = PatternFill(start_color="C6EFCE", fill_type="solid")
                elif decision_msg < 0:
                    res_cell.fill = PatternFill(start_color="FFC7CE", fill_type="solid")
            else:
                res_cell.alignment = Alignment(horizontal="left", vertical="center")
                if any(err in str(decision_msg).upper() for err in ["ERROR", "MISSING", "NOT FOUND"]):
                     res_cell.fill = red_fill

            current_row += 1
            print("----------------------------------------------------------------------------------------------")
        
        ### CHANGE: Apply Merges immediately after each store to prevent cross-store overlap
        for o_no, info in store_order_tracker.items():
            s_row, e_row = info["start"], info["end"]
            if s_row < e_row:
                # Merge Sequence (Col 1)
                ws.merge_cells(start_row=s_row, end_row=e_row, start_column=1, end_column=1)
                ws.cell(row=s_row, column=1).alignment = Alignment(vertical="center", horizontal="center")
                
                # Merge Order # (Col 2)
                ws.merge_cells(start_row=s_row, end_row=e_row, start_column=2, end_column=2)
                ws.cell(row=s_row, column=2).alignment = Alignment(vertical="center", horizontal="right")
                
                # Merge G through K (Cols 7-11: Carrier, Service, Box, Price, Attention)
                for col_idx in range(7, 12):
                    ws.merge_cells(start_row=s_row, end_row=e_row, start_column=col_idx, end_column=col_idx)
                    ws.cell(row=s_row, column=col_idx).alignment = Alignment(vertical="center", horizontal="center")

        current_row += 1 # Blank row between stores

    # Grand Total Savings on last column after last platform
    current_row += 1 
    ws.cell(row=current_row, column=12, value="GRAND TOTAL SAVINGS:").font = Font(bold=True, size=10)
    total_cell = ws.cell(row=current_row, column=13, value=grand_total_savings)
    total_cell.font = Font(bold=True, size=10)
    total_cell.number_format = '"$"#,##0.00'
    total_cell.fill = PatternFill(start_color="FFFF00", fill_type="solid")

    # CREATE DECISION LOG SHEET
    if "Decision Log" in wb.sheetnames:
        del wb["Decision Log"]
    
    log_ws = wb.create_sheet(title="Decision Log")
    log_headers = ["Order #", "SKU", "Shipping DB Cost", "Winner", "Comparison", "Savings", "Decision", "SKU Pkg",
                   "Delivery Time (Days)", "Arrival", "Fallback","LP","Weight","Dims","Shipping Cost","GP","Interchange", "Store Name","Shipping Status"]
    log_ws.append(log_headers)

    for cell in log_ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = black_fill

    green_savings = PatternFill(start_color="C6EFCE", fill_type="solid") # Light Green
    red_savings = PatternFill(start_color="FFC7CE", fill_type="solid")   # Light Red    

    for entry in decision_logs:
        row_data = [
            entry["Order #"], entry["SKU"], entry["DB Cost"], 
            entry["Winner"], entry["Comparison"], entry["Savings"], entry["Decision Type"], 
            entry["Pkg"], entry["Delivery Time"], entry["Arrival"], entry["Fallback"], entry["LP"], 
            entry["Weight"], entry["Dims"], entry["Shipping Cost"], entry["GP"], entry["Interchange"], entry["Store Name"], entry["Shipping Status"]
        ]
        log_ws.append(row_data)

        # Get the row we just added
        curr_log_row = log_ws.max_row
        savings_cell = log_ws.cell(row=curr_log_row, column=5) # Column 5 is Savings

        # Apply Savings Highlighting
        if isinstance(entry["Savings"], (int, float)):
            savings_cell.number_format = '"$"#,##0.00'
            if entry["Savings"] > 0:
                savings_cell.fill = green_savings
            elif entry["Savings"] < 0:
                savings_cell.fill = red_savings

        # Red highlight if Fallback occurred (Column 10 is Status)
        if entry["Fallback"] == "FALLBACK":
            log_ws.cell(row=curr_log_row, column=10).fill = red_fill
        
    # Adjust Column Widths for readability
    for col in log_ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        log_ws.column_dimensions[column].width = max_length + 2

    # List Algorithm
    create_list_algorithm(wb, all_parts_for_list)

    progress_status["percent"] = 100

    wb.save(output_file)

def extract_todays_shipments():
    
    """
        Start of the entire program. Uses Shipstation V1 API to fetch all orders for the day,
        cleans up the SKU names and triggers the Excel writing process (write_grouped_excel).
    """

    start_time = time.perf_counter()

    orders = get_shipments()

    #test_orders = generate_test_orders(2)
    #orders.extend(test_orders)

    store_rows = defaultdict(list)

    for order in orders:
        ship_by = order.get("shipByDate")
        ship_by = order.get("shipByDate")
        
        if ship_by:
            # If it exists, parse it as usual
            ship_by_date = date.fromisoformat(ship_by[:10])
        else:
            # If it's missing, use today's date
            ship_by_date = date.today()
        store_id = order["advancedOptions"].get("storeId")

        for item in order["items"]:
            sku = str(item.get("sku")).strip()
            sku = sku.replace("GMS","MS").replace("AMS","MS").replace("EMS","MS")
            sku = sku.replace("GMK","MK").replace("AMK","MK").replace("EMK","MK")

            # lookup GP# and Interchange # using sku_lookup.py
            record = lookup_sku(sku)
            gp = record["Part #"] if record is not None and "Part #" in record else None
            interchange = record["Interchange (not in order)"] if record is not None and "Interchange (not in order)" in record else None
            
            qty = int(item.get("quantity",1))

            for _ in range(qty):
                store_rows[store_id].append({
                    "Sequence": None,
                    "Order #": order["orderNumber"],
                    "First Name": order.get("shipTo",{}).get("name","").split()[0] if not order.get("First Name") else order.get("First Name"),
                    "Last Name": order.get("shipTo",{}).get("name","").split()[-1] if not order.get("Last Name") else order.get("Last Name"),
                    "SKU": sku,
                    "Part#": gp,
                    "Interchange #": interchange,
                    "Qty": 1,
                    "Carrier": None,
                    "Service": None,
                    "Box": None,
                    "Shipping Price": None,
                    "Attention": None,
                    "Order Date": order["orderDate"][:10],
                    "Ship By": ship_by_date,
                    "State": order["shipTo"]["state"],
                    "Zip":order["shipTo"].get("postalCode"),
                    "Store": order["advancedOptions"].get("storeId")
                })
    if not store_rows:
        return {"status": "empty", "message": "No shipments today"}

    os.makedirs("output", exist_ok=True)
    #output_file = f"output/orders_{today}.xlsx"
    output_file = config.main_file

    write_grouped_excel(store_rows, output_file)

    end_time = time.perf_counter()
    total_seconds = end_time - start_time
    minutes = int(total_seconds // 60)
    seconds = int(total_seconds % 60)
    
    duration = f"{minutes:02d}:{seconds:02d}"

    return {
        "status": "success",
        "rows": len(store_rows),
        "file": output_file,
        "duration": duration
    }

def run_debug_list_algorithm():
    filename = config.main_file
    today_day = str(datetime.today().day)

    try:
        wb = load_workbook(filename)
        if today_day not in wb.sheetnames:
            print(f"Error: Sheet '{today_day}' not found in {filename}")
            return
        
        ws_daily = wb[str(today_day)]

        parts_list = []

        for r in range(3, ws_daily.max_row + 1):
            val = ws_daily.cell(row=r, column=4).value
            # 1. Check if val exists (not None)
            if val is not None:
                # 2. Convert to string and strip whitespace for a clean comparison
                str_val = str(val).strip()
                
                # 3. Check if it's not "nan" (case-insensitive)
                if str_val.lower() != "nan" and str_val != "":
                    parts_list.append(str_val)
        
        create_list_algorithm(wb, parts_list)
        wb.save(filename)
        print(f"Success! Open {filename} to see the result.")
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":

    result = extract_todays_shipments()
    print(result)
