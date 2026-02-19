import base64
from io import BytesIO
from reportlab.lib.pagesizes import LETTER, landscape
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from PyPDF2 import PdfReader, PdfWriter, Transformation
import requests
from datetime import datetime
import config
from openpyxl import load_workbook
import math
from openpyxl.styles import PatternFill
from urllib.parse import quote
import os
import copy
from src.shipstation.rates import V1_SHIPSTATION_API_KEY, V1_SHIPSTATION_API_SECRET, V2_API_KEY
import hashlib
import json
import src.main as main

def get_v1_balance(carrier_code="stamps_com"):
    """Check actual balance via V1 Carriers list."""
    url = "https://ssapi.shipstation.com/carriers"
    response = requests.get(url, auth=(V1_SHIPSTATION_API_KEY, V1_SHIPSTATION_API_SECRET))
    if response.status_code == 200:
        for c in response.json():
            if c.get("code") == carrier_code:
                return c.get("balance", 0.0)
    return 0.0

def merge_labels_to_pdf(metadata_list, output_filename="batch_labels.pdf"):
    """
        Gets the order_metadata_map produced by shipping_label_algo, extracts the data,
        uses create_label_page method and merges all labels into one pdf file.
    """
    writer = PdfWriter()
    for data in metadata_list:
        try:
            page = create_label_page(
                data['base64'], data['name'], data['order_no'], 
                data['address'], data['package'], data['gp_no'], data['interchange'], data['store_name']
            )
            writer.add_page(page)
        except Exception as e:
            print(f"Error processing {data.get('order_no')}: {e}")

    with open(output_filename, "wb") as f:
        writer.write(f)
    return output_filename

def shipping_label_algo(sheet_name):
    BASE_URL = "https://ssapi.shipstation.com"
    SS_AUTH = (V1_SHIPSTATION_API_KEY, V1_SHIPSTATION_API_SECRET)

    ship_balance = get_v1_balance("stamps_com")

    wb = load_workbook(config.main_file)
    ws = wb[sheet_name]

    order_metadata_list = []
    seen_base64_hashes = {}
    created_shipment_ids = []
    session_affected_rows = [] # Track only rows handled in THIS run
    batch_failed = False

    # For the progress bar
    main.progress_status['percent'] = 0
    total_rows = ws.max_row - 1
    processed_count = 0

    try:
        # iterate through rows after header
        for row in range(2, ws.max_row + 1):
            processed_count += 1
            main.progress_status['percent'] = int((processed_count / total_rows) * 100)
            
            # Variable Reset to prevent shipping label mix
            label_json = None
            shipment_id = None
            customer_name = "N/A"
            cust_address = "N/A"

            order_no = ws.cell(row=row, column=1).value
            if not order_no or ws.cell(row=row, column=19).value == "SHIPPED":
                processed_count += 1
                main.progress_status['percent'] = int((processed_count / total_rows) * 100)
                continue

            print(f"Processing Order: {order_no}")

            # Fetch Order from API
            order_response = requests.get(
                f"{BASE_URL}/orders?orderNumber={quote(str(order_no))}", 
                auth=SS_AUTH, 
                timeout=15
            )

            if order_response.status_code != 200:
                print(f"FAILED TO FETCH order {order_no}: {order_response.text}")
                batch_failed = True
                break

            try:
                order_data = order_response.json()
                orders = order_data.get("orders", [])
            except Exception as e:
                print(f"FAILED TO PARSE JSON for {order_no}: {e}")
                batch_failed = True
                break

            if not orders:
                print(f"Order {order_no} not found in ShipStation.")
                continue
                
            matched_order = None
            for o in orders:
                if str(o.get("orderNumber")).strip() == order_no:
                    matched_order = o
                    break
            
            if not matched_order:
                print(f"CRITICAL: API search for {order_no} returned wrong order!")
                batch_failed = True
                break

            order_id = matched_order["orderId"]
            ship_to = matched_order.get("shipTo", {})
            customer_name = ship_to.get("name", "N/A")
            cust_address = f"{ship_to.get('street1')}\n{ship_to.get('city')}, {ship_to.get('state')} {ship_to.get('postalCode')}"

            # --- Data Collection ---
            service_code = ws.cell(row=row, column=7).value
            sku_pkg = ws.cell(row=row, column=8).value
            weight_val = ws.cell(row=row, column=13).value

            # if weight is missing for priority mail, add 1 lbs to it
            if weight_val is None:
                weight_val = 1.0

            dims_str = str(ws.cell(row=row, column=14).value)
            expected_cost = float(ws.cell(row=row, column=15).value)

            # Wallet Check
            if expected_cost > ship_balance:
                print("INSUFFICIENT FUNDS. BUY AND RUN AGAIN. NO ORDERS WERE SHIPPED OR VOIDED")
                batch_failed = True
                break

            ship_balance -= expected_cost

            # Weight/Package Logic
            if service_code == "usps_first_class_mail":
                final_weight = float(weight_val)
            else:
                final_weight = math.ceil(float(weight_val))

            package_code = "package"
            if service_code == "usps_priority_mail":
                package_code = config.pkg_map.get(sku_pkg, "package")
            elif service_code == "usps_first_class_mail":
                sku_pkg = "BAG"

            # Dimensions Logic
            dims = {"units": "inches", "length": 1, "width": 1, "height": 1}

            if sku_pkg in config.DIM_MAP:
                l, w, h = config.DIM_MAP[sku_pkg]
                dims.update({"length": l, "width": w, "height": h})
            else:
                dim_source = None
                if sku_pkg and 'x' in sku_pkg.lower():
                    dim_source = sku_pkg
                elif dims_str and 'x' in dims_str.lower():
                    dim_source = dims_str

                if dim_source:
                    try:
                        d_parts = dim_source.lower().split('x')
                        dims.update({
                            "length": float(d_parts[0]),
                            "width": float(d_parts[1]),
                            "height": float(d_parts[2])
                        })
                    except (ValueError, IndexError):
                        dims.update({"length": 1, "width": 1, "height": 1})
                else:
                    dims.update({"length": 1, "width": 1, "height": 1})

            payload = {
                "orderId": order_id,
                "carrierCode": "stamps_com" if "usps" in service_code else "ups_walleted",
                "serviceCode": service_code,
                "packageCode": package_code,
                "confirmation": "delivery",
                "shipDate": datetime.now().strftime("%Y-%m-%d"),
                "weight": {"value": float(final_weight), "units": "pounds"},
                "dimensions": dims,
                "testLabel": False
            }

            print(f"--- PAYLOAD FOR {order_no} ---")
            print(json.dumps(payload, indent=4)) 
            print("----------------------------")

            # Create Label Request with timeout
            res = requests.post(f"{BASE_URL}/orders/createlabelfororder", auth=SS_AUTH, json=payload, timeout=30)

            if res.status_code == 200:
                label_json = res.json()
                b64_data = label_json.get("labelData")
                actual_cost = float(label_json.get("shipmentCost", 0))
                shipment_id = label_json.get("shipmentId")

                # --- DEBUGGING BLCOK : CHECK FOR DUPLICATE BASE64 ---
                label_hash = hashlib.md5(b64_data.encode()).hexdigest()
                if label_hash in seen_base64_hashes:
                    prev_order = seen_base64_hashes[label_hash]
                    print(f"!!! ALERT: DUPLICATE BASE64 DETECTED !!!")
                    print(f"Order {order_no} (Row {row}) got same label as Order {prev_order}")
                else:
                    seen_base64_hashes[label_hash] = order_no

                # Cost validation logic
                if abs(actual_cost - expected_cost) > 0.01:
                    print(f"COST MISMATCH for {order_no}: Expected {expected_cost}, got {actual_cost}")
                    ws.cell(row=row, column=20).value = actual_cost
                    ws.cell(row=row, column=20).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    # Even if cost mismatches, SS creates the label, so we must add it to the void list
                    created_shipment_ids.append({"shipment_id":shipment_id, "order_no":order_no})
                    batch_failed = True
                    break
                else:
                    created_shipment_ids.append({"shipment_id":shipment_id, "order_no":order_no})
                    session_affected_rows.append(row)
                    
                    order_metadata_list.append({
                        "base64": b64_data,
                        "name": customer_name,
                        "order_no": str(order_no),
                        "address": cust_address,
                        "package": str(sku_pkg),
                        "gp_no": str(ws.cell(row=row, column=16).value or "N/A"),
                        "interchange": str(ws.cell(row=row, column=17).value or "N/A"),
                        "store_name": str(ws.cell(row=row, column=18).value or "N/A")
                    })
                    ws.cell(row=row, column=19).value = "SHIPPED"
            else:
                print(f"API Error for {order_no}: {res.text}")
                batch_failed = True
                break

    except Exception as e:
        # This catches HTTP Connection errors, timeouts, and code crashes
        print(f"CRITICAL ERROR ENCOUNTERED: {e}")
        batch_failed = True

    # --- FINAL CLEANUP / VOIDING LOGIC ---
    if batch_failed:
        print(f"Batch failed. Voiding {len(created_shipment_ids)} labels...")
        # Surgical Rollback
        for row_idx in session_affected_rows:
            ws.cell(row=row_idx, column=19).value = ""

        # Void all labels created in this specific loop
        for item in created_shipment_ids:
            s_id = item["shipment_id"]
            o_no = item["order_no"]
            try:
                void_res = requests.post(f"{BASE_URL}/shipments/voidlabel", auth=SS_AUTH, json={"shipmentId": s_id}, timeout=10)

                if void_res.status_code == 200:
                    void_data = void_res.json()

                    if void_data.get("approved"):
                        print(f" [SUCCESS] Voided Order: {o_no} (Shipment ID: {s_id})")
                    else:
                        print(f"  [WARNING] ShipStation denied void for Order {o_no}: {void_data.get('message')}")
                else:
                    print(f"  [FAILED] HTTP {void_res.status_code} for Order {o_no} - {void_res.text}")

            except Exception as void_err:
                print(f"Could not void Order {o_no} (ID: {s_id}): {void_err}")
        
        wb.save(config.main_file)
        return False
    else:
        # Success path
        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
        if not os.path.exists(downloads_path):
            downloads_path = os.getcwd()

        filename = f"{datetime.now().strftime('%Y-%m-%d_%H%M%S')}_labels.pdf"
        full_path = os.path.join(downloads_path, filename)
        
        output_pdf = merge_labels_to_pdf(order_metadata_list, full_path)
        wb.save(config.main_file)
        return output_pdf
    
def create_label_page(base64_source, name, order_no, address, package, gp_no, interchange, store_name):
    PAGE_W, PAGE_H = landscape(LETTER)
    HALF_W = PAGE_W / 2
    
    # Gap setting (adjust this to make the space bigger or smaller)
    SECTION_GAP = 20 

    # 1. Decode Label
    label_bytes = base64.b64decode(base64_source.split(",")[-1].strip())
    label_reader = PdfReader(BytesIO(label_bytes))
    label_page = label_reader.pages[0]
    mb = label_page.mediabox
    lw, lh = float(mb.width), float(mb.height)

    # 2. Create Layout
    packet = BytesIO()
    can = canvas.Canvas(packet, pagesize=(PAGE_W, PAGE_H))
    
    # --- LEFT SIDE HEADER ---
    header_style = ParagraphStyle(
        'HeaderStyle', 
        fontName='Helvetica-Bold', 
        fontSize=12, 
        leading=14, 
        alignment=TA_CENTER
    )
    header_content = f"{store_name} | {order_no} | {package} | {gp_no}"
    p_header = Paragraph(header_content, header_style)
    
    max_header_w = HALF_W - 40
    w_h, h_h = p_header.wrap(max_header_w, 100)
    p_header.drawOn(can, (HALF_W - w_h) / 2, PAGE_H - 15 - h_h)
    
    # Vertical Divider Line
    can.setStrokeColorRGB(0.8, 0.8, 0.8)
    can.line(HALF_W, 50, HALF_W, PAGE_H - 50)

    # --- RIGHT SIDE PARAMETERS ---
    right_x_anchor = HALF_W + 15 # Moved slightly right for more spacing from line
    max_text_width = (PAGE_W - right_x_anchor) - 30 # Increased margin
    strip_len = PAGE_H - 100
    # The height of each of the two right-side boxes, minus half the gap
    box_height = (strip_len / 2) - (SECTION_GAP / 2)

    # --- A. BOTTOM HALF: SHIPPING INFO (With Gap) ---
    can.saveState()
    path_b = can.beginPath()
    # Box starts at 50, ends before the gap starts
    path_b.rect(HALF_W + 5, 50, HALF_W - 10, box_height)
    can.clipPath(path_b, stroke=0)
    
    can.translate(right_x_anchor, 50 + 10) # +10 for internal bottom padding
    can.rotate(90)
    
    info_style = ParagraphStyle('InfoStyle', fontName='Helvetica', fontSize=13, leading=18, alignment=TA_LEFT)
    info_content = (
        f"<b>SHIPPING INFO</b><br/><br/>"
        f"<b>STORE:</b> {store_name}<br/>"
        f"<b>NAME:</b> {name}<br/>"
        f"<b>ORDER #:</b> {order_no}<br/>"
        f"<b>GP#:</b> {gp_no}<br/>"
        f"<b>INTERCHANGE#:</b> {interchange}<br/>"
        f"<b>ADDRESS:</b> {address.replace('\\n', '<br/>').upper()}"
    )
    
    p_info = Paragraph(info_content, info_style)
    # Available width is box_height minus some padding
    w_i, h_i = p_info.wrap(box_height - 20, max_text_width)
    
    # Draw closer to the center divider
    p_info.drawOn(can, 0, -h_i) 
    can.restoreState()

    # --- B. TOP HALF: PKG (With Gap) ---
    mid_point_start = 50 + box_height + SECTION_GAP # Start after the gap
    
    can.saveState()
    path_t = can.beginPath()
    path_t.rect(HALF_W + 5, mid_point_start, HALF_W - 10, box_height)
    can.clipPath(path_t, stroke=0)

    can.translate(right_x_anchor, mid_point_start)
    can.rotate(90)
    
    pkg_style = ParagraphStyle('PkgStyle', fontName='Helvetica-Bold', fontSize=28, leading=32, alignment=TA_CENTER)

    if package in config.DIM_MAP:
        l,w,h = config.DIM_MAP[package]
        pkg_content = f"PKG: {l}x{w}x{h}<br/>({package})"
    else:
        pkg_content = f"PKG: {package}"
    
    p_pkg = Paragraph(pkg_content, pkg_style)
    w_p, h_p = p_pkg.wrap(box_height - 20, max_text_width)
    
    # Centering within the top box
    p_pkg.drawOn(can, (box_height / 2) - (w_p / 2), -(max_text_width / 2) - (h_p / 2))
    can.restoreState()

    can.save()
    
    # 3. Label Scaling & Merge
    packet.seek(0)
    layout_page = PdfReader(packet).pages[0]
    target_w, target_h = HALF_W - 80, PAGE_H - 140 
    scale = min(target_w / lw, target_h / lh)
    tx = ((HALF_W - (lw * scale)) / 2) - (float(mb.lower_left[0]) * scale)
    ty = ((PAGE_H - (lh * scale)) / 2) - 40 - (float(mb.lower_left[1]) * scale)
    transform = Transformation().scale(scale).translate(tx, ty)
    label_page.add_transformation(transform)
    label_page.mediabox.lower_left = (0, 0)
    label_page.mediabox.upper_right = (PAGE_W, PAGE_H)
    layout_page.merge_page(label_page)
    
    return layout_page